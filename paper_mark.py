#!/usr/bin/env python
# -*- coding:utf-8 -*-
import sys
reload(sys)
sys.setdefaultencoding('utf8')
import sys,os
import re,logging
import time
import commands
import datetime
import openpyxl
import requests
from openpyxl import Workbook
from pyquery import PyQuery

excel_name = "mark_by_tingyun"

file_name = __file__.split('/')[-1].replace(".py","")
#运行过程中的日志文件在执行目录下的lyric_test.log中
logging.basicConfig(level=logging.INFO,
                format='%(asctime)s %(filename)s [line:%(lineno)d] %(levelname)s %(message)s',
                datefmt='%a, %d %b %Y %H:%M:%S',
                filename='%s.log'%file_name,
                filemode='a')

#将日志打印到标准输出（设定在某个级别之上的错误）
console = logging.StreamHandler()
console.setLevel(logging.INFO)
formatter = logging.Formatter('%(name)-12s: %(levelname)-8s %(message)s')
console.setFormatter(formatter)
logging.getLogger('').addHandler(console)


def get_all_pdfs():
    file = commands.getstatusoutput("ls *.pdf")
    if file[0] == 0:
        content = file[1].split("\n")
        return content
    else:
        return None
def get_all_htmls():
    file = commands.getstatusoutput("ls *.html")
    if file[0] == 0:
        content = file[1].split("\n")
        return content
    else:
        return None
'''
def get_text(file_name):
    fp = open("%s"%file_name, "rb")
    parser = PDFParser(fp)
    doc = PDFDocument(parser)
    parser.set_document(doc)
    resource = PDFResourceManager()
    laparam = LAParams()
    device = PDFPageAggregator(resource, laparams=laparam)
    interpreter = PDFPageInterpreter(resource, device)
    for page in PDFPage.create_pages(doc):
        interpreter.process_page(page)
        layout = device.get_result()
        for out in layout:
            if hasattr(out, "get_text"):
                print out.get_text()
def pdf2text(file_name):
    res = commands.getstatusoutput("pdf2txt.py -o %s.html %s"%(re.sub(".pdf","",file_name).replace(" ","\ "),file_name.replace(" ","\ ")))
'''

#如果文件存在不下载，判断html文件，如果html文件都存在，那pdf必然存在
#下载到./pdf_file/，将pdf转化成对应的html存在./html_file/路径下 , 返回pdf_html的名字
def download_pdf(url):
    pdf_name = re.search("\w*$", url).group().replace(" ", "")
    isexists = os.path.exists("./html_file/%s.html"%pdf_name)
    if isexists:
        logging.info("html文件存在，略过....")
        return pdf_name
    else:
        response = requests.get(url)
        if re.findall("[4|5|3]", str(response.status_code)):
            logging.warning("the url %s is failed ,status code is %s !" % (link, response.status_code))
            return None
        else:
            content = response.content
            with open("./pdf_file/%s.pdf"%pdf_name,"wb") as f:
                f.write(content)
                f.close()
            a = commands.getstatusoutput("pdf2htmlEX ./pdf_file/%s.pdf ./html_file/%s.html"%(pdf_name,pdf_name))
            if a[0] == 0:
                logging.info("转换成html成功....")
                return pdf_name
            else:
                logging.info("html转换失败....")

def loadhtml_pyquery(file_name):
    res = ''
    with open("%s"%file_name,'rb') as f:
        content = f.read()
        res = PyQuery(content)
        f.close()
    return res

def loadurl_pyquery(url):
    response = requests.get(url)
    if re.findall("[4|5|3]", str(response.status_code)):
        logging.warning("the url %s is failed ,status code is %s !" % (link, response.status_code))
        return None
    else:
        content = response.content
        py_content = PyQuery(content)
        return py_content

#使用旧的版本pdf2text转换成的html的解析方法，现已废弃
def old_getinfo_fromhtml(file_name):
    py_content = loadhtml_pyquery(file_name)
    for i in py_content('div').items():
        for j in i('span').items():
            if re.findall("email",j.text()):
                temp = j.text()
                email_list = re.findall("\(.*?\)",temp)
                res = []
                for i in email_list:
                    res.append(i.replace("(","").replace(")",""))
                #names = py_cpntent("div:contains('OPEN')").next().text()
                #print names


def init_excel(excel_name):
    exists = commands.getstatusoutput("ls %s.xlsx"%excel_name)
    if exists[0] == 0: #存在,返回文件handler
        excel = openpyxl.load_workbook("%s.xlsx"%excel_name)
        return excel
    else:   #创建新的excel
        excel = Workbook()
        sheet = excel.active
        excel.create_sheet(index=0,title="mark_by_tingyun")
        wb1 = excel.get_sheet_by_name("mark_by_tingyun")
        wb1['A1'] = "序号"
        wb1['B1'] = "title"
        wb1['C1'] = "Author"
        wb1['D1'] = "Email"
        wb1['E1'] = "Organization"
        wb1['F1'] = "Link"
        wb1['G1'] = "Abstract"
        wb1['H1'] = "Keywords"
        wb1['I1'] = "Fields"
        wb1['J1'] = "Acknowlegement"
        wb1['K1'] = "Publication_Date"
        wb1['L1'] = "Publisher"
        wb1['M1'] = "journal_or_conference"

        excel.save("%s.xlsx"%excel_name)
        logging.info("excel文件 %s.xlsx 创建成功"%excel_name)
        excel = openpyxl.load_workbook("%s.xlsx" % excel_name)
        return excel

def getinfo_from_html(file_name):
    info = {}
    py_content = loadurl_pyquery(b[0])
    #y2和y3是标题
    title = py_content(".y2").text() +" " + py_content(".y3").text()

    #y4和y5是人物和标注对应关系，并且是轮着来的，先y4后y5，再y4 ; but 我们切出来的每个div都带有空格，所以可以按照空格切分，再分别匹配
    author_name = py_content(".y4").text()
    author_dig = py_content(".y5").text().split(" ")
    r = re.compile("[,&]")
    author_name = r.split(author_name)
    author_group = {}
    if len(author_name) == len(author_dig):
        for i in range(len(author_name)):
            author_group[author_name[i]] = author_dig[i].split(",")

    email = {}
    email_name = re.findall("(?<=ddressed to).*\)", py_content(".ff6.fs5").text())[0].split("or")
    for i in email_name:
        name_abb = re.search(".*(?=\()", i).group().replace(".", "").replace("to", "").replace(" ", "")
        e_name = re.search("\(.*\)", i).group().replace("(email:", "").replace(")", "")
        email[e_name] = name_abb

    #提出每个author的大写字母，拼起来即可，英文名字首字母是大写 | 还有带有 - 也得找到
    author2email = {}   #得出作者对应的邮箱,格式为：{'name':'email1;email2;....'}
    #res = ""
    for i in author_name:
        for j,v in enumerate(email.values()):
            res = ""
            temp = ''.join(re.findall("[A-Z|-]", i))
            if temp == v:
                res += email.keys()[j]+";"
            if res:
                author2email[i] = re.sub(";$","",res)
    noEmail_author = list(set(author_name).difference(author2email.keys()))
    for i in noEmail_author:
        author2email[i] = "np.nan"

    #ff6为所有的所在集团标注，其中fs7为数字，fs5为集团名称    一一对应的关系
    #取得和dig一样长度的name，多余的是邮箱地址，这个需要根据人名来判断对应关系
    group = {}      #格式为 {'group_dig':'group_name'}
    group_dig = py_content(".ff6.fs7").text().split(" ")
    #希望所有的格式都和这个一样，是在标注后面加上 Correspondence....
    group_name = re.sub("\. Correspondence.*\)","",py_content(".ff6.fs5").text()).split(".")
    if len(group_dig) == len(group_name):
        for i in range(len(group_dig)):
            group[group_dig[i]]=group_name[i]

    author2group = {}   #多个所属，用;隔开即可，格式为：{'name':'group1;\ngroup2;\n....'}
    for i in author_group.keys():
        temp = ';\n'.join([group[x] for x in author_group[i]])
        author2group[i] = temp
    #论文的摘要
    abstract = py_content(".ff1.fs4").text()

    #Acknowledgements
    acknow = re.search("(?<=Acknowledgements).*(?=Author contributions)",py_content.text()).group()

    info['abstract'] = abstract
    info['acknow'] = acknow
    info['author2email'] = author2email #author -- email ,email是一个以 ；\n 分割的字符串  ,在写到excel的时候填充字符串即可
    info['author2group'] = author2group #author -- group ,group是一个以 ；\n 分割的字符串  ,在写到excel的时候填充字符串即可

    #以上值不存在的时候，均是为"np.nan"
    return info

def get_info_from_web(link):
    info = []
    # url_head = re.search("http://.*?(?=/)",link).group()
    py_content = loadurl_pyquery(link)

    #以上的匹配都无法正确找到acknow，这里采用读取pdf，再转换
    #return
    title = py_content("h1").text()
    abstract = py_content('.abstractSection p ').text()
    author = [x.text() for x in list(py_content('.authors:first .entryAuthor').items())]

    pdf_url = re.sub("doi", "doi/pdf", link)
    pdf_name = download_pdf(pdf_url)
    html_content = loadhtml_pyquery("./html_file/%s.html" % pdf_name)
    # 在网页中无法找到这个Acknowledgments
    if re.findall("Acknowledgments",html_content.text()):
        acknow = re.findall("(?<=Acknowledgments).*(?=References)", html_content.text())
    else:
        acknow = "np.nan"

    fields = "np.nan"
    keywords = "np.nan"
    email = "np.nan"
    belong = "np.nan"
    link = link
    date = re.search("(?<=Online ).*(?=doi:)", py_content('.articleInfo').text()).group()
    publisher = re.search("(?<=under a ).+(?=\. For)", py_content('.OATextContainer > p').text()).group()

    for i in author:
        #从下表0开始： 作者 0，标题 1，摘要 2，致谢 3，领域 4，关键词5
        temp = [str(i) , str(title) , str(abstract) , str(acknow) , fields , keywords , email , belong ,link ,date , publisher]
        info.append(temp)
    return info

def write_excel(url):
    wb = init_excel(excel_name)
    wb1 = wb.get_sheet_by_name("mark_by_tingyun")
    try:
        info = get_info_from_web(url)
    except Exception, e:
        info = None
        logging.error("啥都没有 %s"%url)
    if info:
        for i in info:
            wb1.append(['',i[1],i[0],i[6],
                      i[7],i[8],i[2],i[5],i[4],
                      i[3],i[9],i[10],''])
        #wb1.append(['','','','','','','','','','','','','','','',''])
        #每个表添加一个空数据行
        wb.save("%s.xlsx" % excel_name)
    else:
        pass


if __name__=='__main__':
    a = get_all_pdfs()
    b = get_all_htmls()
    #getinfo_from_html(b[0])
    url = "http://www.mitpressjournals.org/doi/10.1162/COLI_a_00132"
    write_excel(url)



